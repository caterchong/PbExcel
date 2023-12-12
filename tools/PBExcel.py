# -*- coding: utf-8 -*-
# https://googleapis.dev/python/protobuf/latest/index.html
from google.protobuf.descriptor_pb2 import FileDescriptorSet
from google.protobuf import json_format
from google.protobuf.json_format import MessageToDict
from google.protobuf.descriptor_pool import DescriptorPool
from google.protobuf import message_factory
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl import load_workbook
import platform,datetime,json, argparse,subprocess,codecs,time,os,sys
import Config

# protoc --descriptor_set_out=test.proto.desc --include_source_info  test.proto  生成二进制文件
# https://developers.google.com/protocol-buffers/docs/reference/cpp/google.protobuf.descriptor.pb
# 这里面可以看到具体的descriptor描述
def GenMetaFromDesc(desc):
  #print(text_format.MessageToString(desc))
  dict_obj = MessageToDict(desc)
  proto_obj = dict_obj["file"][0]

  ########  fill comments tag to  proto info ######
  commentList = [x for x in proto_obj["sourceCodeInfo"]["location"] if "leadingComments" in x and (x["path"][0] == 5 or x["path"][0] == 4)]
  for comment in commentList:
    # 去除掉多余的空格 tab等等
    tagList = [x for x in comment["leadingComments"].split("\n") if x.startswith("@")]
    tagMap = {}
    for tag in tagList:
      tagName = tag.split()[0].replace("@","")
      if len(tag.split()) > 1:
        tagValue = tag.replace("@"+tagName, "").strip()
      else:
        tagValue = True
      tagMap[tagName] = tagValue

    if comment["path"][0] == 5 and len(comment["path"]) == 4 and comment["path"][2] == 2:#repeated EnumDescriptorProto enum_type = 5;   // 所有的枚举(enum)类型定义
      proto_obj["enumType"][comment["path"][1]]["value"][comment["path"][3]]["tag"] = tagMap
    elif comment["path"][0] == 4:#repeated DescriptorProto message_type = 4;    // 所有的消息(message)类型定义
      if len(comment["path"]) == 2: # comment for message
        proto_obj["messageType"][comment["path"][1]]["tag"] = tagMap
      elif  len(comment["path"]) == 4: # comment for message entry
        proto_obj["messageType"][comment["path"][1]]["field"][comment["path"][3]]["tag"] = tagMap
  return proto_obj

# 为某个entry生成注释
def GenCommentFromMeta(meta):
  return json.dumps(meta, sort_keys=True, indent=4, ensure_ascii=False)

def GetNameFromMeta(meta):
  if "tag" in meta and "name" in meta["tag"]:
    return meta["tag"]["name"]
  return meta["name"]

def GenCellFromFieldMeta(protoDesc, columnList, prefix, fieldMeta):
  columeName = prefix + GetNameFromMeta(fieldMeta)
  if fieldMeta["type"] == "TYPE_MESSAGE":
    if fieldMeta["label"] == "LABEL_OPTIONAL":
      FlatMessage(protoDesc, columnList, columeName+".", fieldMeta["typeName"].split(".")[-1])  # typeName 可能是 .pbjson.SubMsg
    else:
      FlatMessage(protoDesc, columnList, columeName+"[1].", fieldMeta["typeName"].split(".")[-1])
  else:
    if fieldMeta["label"] == "LABEL_OPTIONAL":
      cell=WriteOnlyCell(None, columeName)
      cell.comment = Comment(GenCommentFromMeta(fieldMeta),author=None,height=200, width=400)
      columnList.append(cell)
    else:
      cell=WriteOnlyCell(None, columeName)   # todo
      cell.comment = Comment(GenCommentFromMeta(fieldMeta),author=None,height=200, width=400)
      columnList.append(cell)
    
# 解决嵌套消息的展开问题
def FlatMessage(protoDesc, columnList, prefix, messageName):
  meta = [x for x in protoDesc["messageType"] if x["name"] == messageName][0]
  [GenCellFromFieldMeta(protoDesc, columnList, prefix, x) for x in meta["field"]]

def GetExcelMeta(meta):
  excelMessage = [x for x in meta["messageType"] if "tag" in x and ("excel" in x["tag"])]
  if len(excelMessage) != 1:
    raise RuntimeError("defination wrong, should only have one message taged with @excel, current cnt {}".format(len(excelMessage)))
  return excelMessage[0]

def GenExcel(protoFile):
  desc = ConvertProtoToFileDescriptor(protoFile)
  meta = GenMetaFromDesc(desc)

  if not protoFile in Config.Config:
      raise RuntimeError("ERROR, proto file {} not found in Config.py".format(protoFile))
  
  excelName = Config.Config[protoFile]["excel"]

  if os.path.exists(excelName):
    raise RuntimeError("ERROR, proto file {} already exist, if wannna generating a new one, pls rename existing file".format(excelName))

  excelMessage = GetExcelMeta(meta)
  wb = Workbook()
  wb.remove(wb.active)
  
  for field in excelMessage["field"]:
    if field["type"] != "TYPE_MESSAGE":
      raise RuntimeError("ERROR, fields [{}] type must be message".format(field["name"]))
    worksheet = wb.create_sheet(GetNameFromMeta(field))
    row=[]
    FlatMessage(meta, row, "", field["typeName"].split(".")[-1])
    worksheet.append(row)
  wb.save(excelName)

# 用递归代码复用
def ConvertSingleTagName(meta, messageName, tagNameSplitList):
  #print("tagNameSplitList", tagNameSplitList)
  if not tagNameSplitList:   # empty, return
    return []
  messageMeta = [x for x in meta["messageType"] if x["name"] == messageName.split(".")[-1]]
  if not messageMeta:  # size 0, meta not found
    return []
  
  ans=[]
  entryMeta=[x for x in messageMeta[0]["field"] if GetNameFromMeta(x) == tagNameSplitList[0]]
  #print("entryMeta:  ", entryMeta)
  if entryMeta:
    ans.append(entryMeta[0])   #直接换meta是为了方便查找tag，做一些处理
    if len(tagNameSplitList)>1 and tagNameSplitList[1].isdigit():
      ans.append(tagNameSplitList[1])
    if len(tagNameSplitList) >len(ans): # still have sub message
      ans += ConvertSingleTagName(meta, entryMeta[0]["typeName"], tagNameSplitList[len(ans):])
    return ans
  else:
    raise RuntimeError("entryName {} not found".format(tagNameSplitList[0]))

# allMeta["enumType"]示例: [{'name': 'Gender', 'value': [{'name': 'GENDER_MALE', 'number': 1, 'tag': {'name': '男'}}, {'name': 'GENDER_FEMALE', 'number': 2, 'tag': {'name': '女'}}]}]
def GetEnumValue(allMeta, enumTypeName, value):
  for enumMeta in allMeta["enumType"]:
    if enumMeta["name"] == enumTypeName:
      for singleEnum in enumMeta["value"]:
        if value == GetNameFromMeta(singleEnum):
          return singleEnum["number"]

  raise RuntimeError("enum value invalid, type:{}, value{}".format(enumTypeName, value))

DefaultValueMap={
  "TYPE_BOOL":False,
  "TYPE_STRING":"",
  "TYPE_INT32": 0,
  "TYPE_UINT32": 0,
  "TYPE_INT64": 0,
  "TYPE_UINT64": 0,
  "TYPE_FLOAT": 0,
}
  

# 只转换一个，origValue只是为了debug打印使用
def SingleFinal(allMeta, meta, row, col, origValue, singleOrigValue):
  if "tag" in meta and "datetime" in meta["tag"]:
      dt = datetime.datetime.strptime(singleOrigValue, "%Y-%m-%d %H:%M:%S")
      utc_time = time.mktime(dt.timetuple())
      return int(utc_time)

  entryType = meta["type"]
  if entryType == "TYPE_STRING":
    return str(singleOrigValue)
  if entryType =="TYPE_BOOL":
    if singleOrigValue  == "是" or singleOrigValue==1:
      return True
    elif singleOrigValue  == "否" or singleOrigValue==0 or singleOrigValue=="":
      return False
    else:
      raise RuntimeError("entryName {},row{}, col{}, origValue {} invalid, 必须为是/否/0/1".format(meta["name"], row, col, origValue))
  
  if entryType in ["TYPE_INT32","TYPE_UINT32", "TYPE_INT64", "TYPE_UINT64"]:
      return int(singleOrigValue)
      #raise RuntimeError("entryName {},type {}, row {}, col {}, origValue {} invalid".format(meta["name"], entryType, row, col, origValue))

  if entryType =="TYPE_FLOAT":
    return float(singleOrigValue)
  if entryType == "TYPE_ENUM": # entryMeta示例: {'name': 'gender', 'number': 6, 'label': 'LABEL_OPTIONAL', 'type': 'TYPE_ENUM', 'typeName': '.pbjson.Gender', 'jsonName': 'gender'}
      return GetEnumValue(allMeta, meta["typeName"].split(".")[-1], singleOrigValue)
  
  raise RuntimeError("entryName{},row{}, col{}, origValue{} ,type {} not support".format(meta["name"], row, col, origValue, entryType))


def GetFinalVal(allMeta, meta, row,col, value):
  if value == None and meta["label"] == "LABEL_REQUIRED":
    return DefaultValueMap[meta["type"]]
  if value == None:
    return
  #print(meta)
  if meta["label"] == "LABEL_REPEATED":
    if value == None:
      return
    valueList = str(value).split(";")
    finalList=[SingleFinal(allMeta, meta, row, col, value, v.strip()) for v in valueList]  #去掉前后多余的空格
    return finalList
  else:
    return SingleFinal(allMeta, meta, row, col, value, value)

# 填充数据到mapData
def FillData(allMeta, root, headerSplit,row,col, value,filter):
  if not headerSplit:
    return
  if filter(headerSplit[0]):
    return
  entryType = headerSplit[0]["type"]
  entryName = headerSplit[0]["name"]
  entryLabel = headerSplit[0]["label"]

  if entryType == "TYPE_MESSAGE":
    if value == None:
      return
    if entryLabel == "LABEL_OPTIONAL":
      root[entryName] = {} if entryName not in root else root[entryName]
      FillData(allMeta, root[entryName], headerSplit[1:],row,col, value,filter)
    else:  # repeated, should have a index, index start from 1
      index = int(headerSplit[1])-1
      root[entryName] = [] if entryName not in root else root[entryName]
      if index > len(root[entryName]):
        raise RuntimeError("entryName {} index {}>{}, too big".format(entryName, index, len(root[entryName])))
      if index == len(root[entryName]):
        root[entryName].append({})
      FillData(allMeta, root[entryName][index], headerSplit[2:],row, col, value,filter)
  else:  # basic 
    try:
      convertValue = GetFinalVal(allMeta, headerSplit[0],row,col, value)
      if convertValue != None:
        root[entryName] = convertValue
    except Exception as e:
      print("ERROR!!!   header:{}, row:{},col,{}, value:{}".format(headerSplit[0]["name"],row+1,col+1, value))
      raise(e)


# headerRow 列标题所在的行， startCol：数据开始的列，这两个参数是为了应对客户端特殊的Excel格式，计数都是从1开始
# filter 是否需要过滤掉某个sheet, 某个member
def GenPythonObj(meta, excelFile, filter, headerRow=1, startCol = 1):
  if excelFile == None:
    raise RuntimeError("ERROR, excelFile or outputFile None")
  excelMessage = GetExcelMeta(meta)
  
  root = {}
  wb = load_workbook(filename = excelFile, data_only=True)
  for sheet in wb.worksheets:
    if sheet.title not in [GetNameFromMeta(fieldMeta) for fieldMeta in excelMessage["field"]]:
      print("\twarning: {} not in message defination, ignore".format(sheet.title))
      continue
    fieldMeta = [fieldMeta for fieldMeta in excelMessage["field"] if sheet.title == GetNameFromMeta(fieldMeta)][0]
    print(fieldMeta)
    if filter(fieldMeta):
      continue

    print("Excel:",excelFile, "Sheet:", sheet.title, "MaxRow:", sheet.max_row, "MaxCol:",sheet.max_column)
    headerText=[ cell.value for cell in sheet[headerRow]]  #第一行

    #headerSplit = [ConvertSingleTagName(meta,  fieldMeta["typeName"], [] if x == None else x.replace("[", ".").replace("]", "").split(".")) for x in headerText]
    
    headerSplit=[]
    for x in headerText:
      textSplit=[]
      if x != None:
        textSplit= x.replace("[", ".").replace("]", "").split(".")   # a.b 拆分成[a b], a[1].b拆分成[a 1 b]
      headerSplit.append(ConvertSingleTagName(meta, fieldMeta["typeName"], textSplit))
    # headerSplit最终变成[ a_entryMeta 1  b_entry_meta], 方便后面的处理

    #print("fieldMeta", fieldMeta)
    if fieldMeta["label"] == "LABEL_REPEATED": #  一个sheet只配置多行数据
      root[fieldMeta["name"]]=[]
      for row in range(headerRow, sheet.max_row):  #ignore header
        singleRow = {}
        for col in range(startCol -1, min(sheet.max_column,len(headerText))):  # 规避有的时候有些excel在一些空白的地方乱写数据
          cell = sheet.cell(row=row+1, column=col+1)  # 行列都从1开始
          FillData(meta, singleRow, headerSplit[col], row,col, cell.value,filter)
        #print(singleRow)
        if singleRow:#如果是空的，就算了
          root[fieldMeta["name"]].append(singleRow)
    else:     #  一个sheet只配置一行数据, 这个时候就没有必要生成array了
      singleRow = {}
      for col in range(startCol-1, min(sheet.max_column,len(headerText))):
        cell = sheet.cell(row=2, column=col+1)  # 行列都从1开始
        FillData(meta, singleRow, headerSplit[col], 1,col, cell.value,filter)
      #print(singleRow)
      if singleRow:
        root[fieldMeta["name"]]=singleRow
  return root

def serverFilter(meta):
  if "tag" in meta and "client" in meta["tag"]:
    return True
  return False

def clientFilter(meta):
  if "tag" in meta and "server" in meta["tag"]:
    return True
  return False

def GenCfg(protoFile):
  if not protoFile in Config.Config:
      raise RuntimeError("ERROR, proto file {} not found in Config.py".format(protoFile))

  excelFile = Config.Config[protoFile]["excel"]

  desc = ConvertProtoToFileDescriptor(protoFile)
  meta = GenMetaFromDesc(desc)

  if excelFile == None:
    raise RuntimeError("ERROR, excelFile not configured")

  excelMessage = GetExcelMeta(meta)


  if "serverCfg" in Config.Config[protoFile]:
    pythonObj = GenPythonObj(meta, excelFile, serverFilter)
    outputFile = Config.Config[protoFile]["serverCfg"]
    content = json.dumps(pythonObj, sort_keys=False, indent=4, separators=(",", ":"), ensure_ascii=False)
    codecs.open(outputFile, "w", encoding="utf-8").write(content)
    print("serverCfg " + outputFile)

  if "clientCfg" in Config.Config[protoFile]:
    outputFile = Config.Config[protoFile]["clientCfg"]
    pythonObj = GenPythonObj(meta, excelFile, clientFilter)
    content = json.dumps(pythonObj, sort_keys=False, indent=4, separators=(",", ":"), ensure_ascii=False)
    codecs.open(outputFile, "w", encoding="utf-8").write(content)
    print("clientCfg " + outputFile)

# convert from proto text to file decriptor binary string
def ConvertProtoToFileDescriptor(protoFile):
  toolPath=os.path.dirname(__file__)
  cmd="{}/protoc.exe --descriptor_set_out={}.desc --include_source_info  {}".format(toolPath,protoFile, protoFile)
  if platform.system() == "Windows":
    pass
  else:
    cmd="{}/protoc --descriptor_set_out={}.desc --include_source_info  {}".format(toolPath,protoFile, protoFile)

  subprocess.check_output(cmd, shell=True)
  with open("{}.desc".format(protoFile), "rb") as f:
    desc = FileDescriptorSet.FromString(f.read())
  os.remove("{}.desc".format(protoFile))

  return desc

if __name__ == '__main__':
  parser = argparse.ArgumentParser(description='PBExcel')

  parser.add_argument('-p', '--proto', dest='proto', action='store',
                    help='protobuf file name')
  parser.add_argument('-a', '--action', dest='action', action='store', required=True,
                    help='action list: GenExcel GenCfg GenAll')
  #parser.add_argument('-o', '--output', dest='output', action='store', required=True,
  #                  help='output file name')
  #parser.add_argument('-e', '--excel', dest='excel', action='store',
  #                  help='excel file')
  #parser.add_argument('-f', '--format', dest='format', action='store', default="json",
  #                  help='output format, valid when action is GenCfg, can be lua/json/pb')
  args = parser.parse_args()

  if args.action == "GenExcel":
    GenExcel(args.proto)
  elif args.action == "GenCfg":
    GenCfg(args.proto)
  elif args.action == "GenAll":
    print(Config.Config.keys())
    for proto in Config.Config.keys():
      GenCfg(proto)
  else:
    print("invalid action")
    exit(1)
